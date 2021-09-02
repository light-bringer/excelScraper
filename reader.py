import json
from openpyxl import load_workbook


class Excel:
    @staticmethod
    def readFromExcel(xls_path):
        workbook = load_workbook(xls_path)
        sheet = workbook.active
        products = []
        for row in sheet.iter_rows(min_row=2,
                                   min_col=0,
                                   max_col=6,
                                   values_only=True):
            product = {
                "sku": row[0],
                "asin": row[1],
                "product_name": row[2],
                "keyword1": row[3],
                "keyword2": row[4],
                "keyword3": row[5],
            }
            products.append(product)

        return products


# class CSV:
#     @staticmethod
#     def readFromCSV(csv_file):
#

def read_xls_csv(filepath):
    if filepath.split('.')[-1] == 'xls' or 'xlsx':
        return Excel.readFromExcel(filepath)
    else:
        print("Method not implemented")
        # TODO : Implement for CSV
        pass
